# -*- coding: utf-8 -*-
"""

=================================
Author: keen
Created on: 2019/9/2

E-mail:keen2020@outlook.com

=================================


"""

import openpyxl


class Case(object):
    # 用这个类来存储用例
    def __init__(self, attrs):
        """
        初始化用例
        :param attrs: zip类型--> [（key1,value1),（key2,value2)....]
        """
        for item in attrs:
            setattr(self, item[0], item[1])


class ReadExcel(object):
    """
    读取Excel数据
    """

    def __init__(self, file_name, sheet_name):
        """
        初始化读取对象
        :param file_name: 文件名，测试用例文件--> str
        :param sheet_name: 表单名--> str
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        self.wb = openpyxl.load_workbook(self.file_name)  # 打开工作簿，传入的指定文件名
        self.sheet = self.wb[self.sheet_name]    # 选取表单，传入的指定表单

    def close(self):
        self.wb.close()

    def read_line_data(self):
        """
        执行读取数据
        :return: list
        """
        # 打开工作簿
        self.open()
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)

        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            # 对title是否为空进行过滤，容错机制
            if title.value:
                titles.append(title.value)

        # 定义一个空列表用来存储所有的用例
        cases = []
        # 从第2行开始，就是测试用例数据了
        for case in rows_data[1:]:
            # data用例临时存放每行的用例数据
            data = []
            for cell in case:
                data.append(cell.value)
            # 将title与测试用例数据组合，形成每行测试用例,一行行读取数据，无需加list
            case_data = zip(titles, data)
            # 创建一个Case类的对象，用来保存用例数据，
            case_obj = Case(case_data)
            # 将该条数据放入cases中
            cases.append(case_obj)
        # 关闭工作簿,并返回读取的数据-->list
        self.close()
        return cases

    def read_column_data(self, read_column):
        """
        执行读取数据
        :param read_column:指定读取的列[1,2,3,4...]
        :return: 返回一个list，每个元素为一个用例,dict
        """
        # 打开工作簿
        self.open()
        if len(read_column) == 0:
            return self.read_line_data()

        max_r = self.sheet.max_row
        cases = []  # 存储所有用例
        titles = []  # 存放标题
        for row in range(1, max_r+1):
            if row != 1:    # 排除表头，获取数据
                case_data = []      # 存储该行的数据
                for column in read_column:
                    case_rc = self.sheet.cell(row, column).value   # 获取指定单元格的数据
                    case_data.append(case_rc)
                case = zip(titles, case_data)  # 将表头和数据进行打包，--> 第1次遍历已经添加了表头,一行行读取数据，故无需加list
                case_obj = Case(case)
                cases.append(case_obj)
            else:   # 获取表头信息
                for column in read_column:
                    title = self.sheet.cell(row, column).value
                    if title:
                        titles.append(title)
        # 关闭工作簿
        self.close()
        return cases

    def write_data(self, row, column, value):
        # 打开工作簿
        self.open()
        # 指定位置写入数据
        self.sheet.cell(row=row, column=column, value=value)
        # 保存数据
        self.wb.save(self.file_name)
        # 关闭工作簿
        self.close()


if __name__ == '__main__':

    import os
    from common.constant import DATA_DIR

    wb = ReadExcel(os.path.join(DATA_DIR, 'api_automation_course.xlsx'), "login")
    cases = wb.read_line_data()
    for case in cases:
        print(case.title)
















